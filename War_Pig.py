'''
This is a bot that allows for the users to create coordination channels
for war purposes. It can either do that through bulk creating channels
with a CSV as input or create a single one.

:Date: 7/4/2020
:Versions: 1
:Author:
    -Piggy
'''

#heroku local:run python war_pig.py
import csv
import re
import discord
import requests
#import excel2img
import time
import numpy as np
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
import math
import gspread
import asyncio
from collections import OrderedDict
from scipy import stats
from bs4 import BeautifulSoup
from discord.ext import commands
from openpyxl.styles import Font
from openpyxl.styles import colors
from API import get_pnw_name
from API import get_pnw_mil
from API import get_war_info
from API import get_all_war_info
from API import get_leader
from API import get_cities
from API import req_info
from API import ID_info
from openpyxl import load_workbook
from typing import Optional
import os
import sys

# Starting up Discord
intents = discord.Intents.default()
intents.members = True
client = commands.Bot(command_prefix = '!', intents=intents)
client.remove_command('help')

gaccount = gspread.service_account(filename = 'google-credentials.json')

#gsheet = gaccount.open("Discord Tracking Sheet").sheet1
# make it [1:]
'''member_names = gsheet.col_values(1)
member_names.pop(0)
nation_id = gsheet.col_values(2)
nation_id.pop(0)
dis_id = gsheet.col_values(4)
dis_id.pop(0)
nation_id = list(map(int, nation_id))
dis_id = list(map(int, dis_id))
member_dict = dict(zip(member_names, dis_id))
nation_dict = dict(zip(dis_id, nation_id))  
rev_nation_dict = dict(zip(nation_id, dis_id))'''

#sham_api_key = S3Connection(os.environ['ShamAPIKey'])
#sham_ip = S3Connection(os.environ['ShamAPIIP'])

sham_api_key = os.environ.get('ShamAPIKey')
sham_ip = os.environ.get('ShamAPIIP')

# Setting up the database
membership_db = requests.get(f'http://{sham_ip}:8080/discord/?key={sham_api_key}&alliance=5049').json()
member_names = [member['leader'] for member in membership_db]
dis_id = [int(member['DiscordID']) for member in membership_db]
nation_id = [int(member['_id']) for member in membership_db]
member_dict = dict(zip(member_names, dis_id))
nation_dict = dict(zip(dis_id, nation_id))

# Setting up the spheres sheet
wargsheet = gaccount.open("Carthago Milcom & Personnel").worksheet("Spheres")
sphere_names = [sphere.lower() for sphere in wargsheet.col_values(1)[1:]]
sphere_alliances = [sphere.split(',') for sphere in wargsheet.col_values(3)[1:]]
spheres = dict(zip(sphere_names, sphere_alliances))

warmembergsheet = gaccount.open("Carthago Milcom & Personnel").worksheet("Member Info")

category_list = ['[CANNAE BUT COUNTER]', '[CANNAE BUT COUNTER 2]', '[CANNAE BUT COUNTER 3]', '[BARRACKS]', '[BARRACKS 2]', '[BARRACKS 3]']

# Getting the variables from Heroku


@client.event
async def on_ready():
    '''
    Lets user know bot is ready
    '''
    print('Ready to go')
    await client.change_presence(status=discord.Status.do_not_disturb, activity=discord.Game("Destroying mad cash"))



war_types = ["RAID", "ORDINARY", "ATTRITION"]
@client.command()
async def bulk_create(ctx, api: Optional[str] = "pnw", war_type: Optional[int] = 0):
    '''
    Creates a list of channels based on csv target list
    '''
    category_list_id = []

    for category in category_list:
        category_list_id.append(discord.utils.get(ctx.guild.categories, name = category))

    #Sees if the user has permissions to manage channels in the category and access bot
    category_list_id = list(filter(None, category_list_id))

    if any(category.permissions_for(ctx.author).manage_channels for category in category_list_id):
        #for loop to get attachment
        for attachment in ctx.message.attachments:
            await attachment.save(f'csv/{attachment.filename}')
            await ctx.send('Creating war channels...')
            break
        #If no attachment is found
        else:
            await ctx.send('No Attachment Found')

        #Access attachment and creates channels with it
        if api == 'pnw':
            with open(f'csv/{attachment.filename}') as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=',')
                line = 0
                update_dict()
                #Goes through each row to create channels
                for row in csv_reader:
                    #Makes sure this is not the heading and it isn't an invalid row before creating a channel
                    if line != 0 and row[1] != '#ERROR!':
                        print(f'{row[1]} is the target. Attacker 1 is {row[4]}, attacker 2 is {row[6]},\
                         and attacker 3 is {row[8]}')
                        target_name = row[1]
                        target_id = row[2]
                        attackers = [row[4], row[6], row[8]]
                        defenders = [row[11],row[13],row[15],row[17],row[19]]
                        channel_name = target_name.replace(' ', '-') + '-' + target_id

                        #Creates the channel and edits the topic
                        channel = None
                        for category in category_list_id:
                            try: 
                                print(row[0])
                                channel = await ctx.guild.create_text_channel(channel_name, category = category, topic = f'War on {row[0]}')
                                break
                            except:
                                pass

                        war_embed = discord.Embed(title= f"⚔️ __Target: {' '.join(channel_name.split('-')[:-1])}__", 
                            description= f"Please declare {war_types[war_type]} war on {row[0]}", color=0xcb2400,
                            url = f'https://politicsandwar.com/nation/war/declare/id={row[0].split("=")[1]}')


                        mil_count = get_pnw_mil(f'https://politicsandwar.com/nation/id={row[0].split("=")[1]}')

                        war_embed.add_field(name = '__Military Information:__', value = f'{channel_name.split("-")[0]} has {mil_count["Soldiers"]} soldiers, {mil_count["Tanks"]} tanks, {mil_count["Planes"]} planes, and {mil_count["Ships"]} ships', inline = False)

                        ping_list = await coord_perms(attackers, channel, channel_name, ctx)
                        #Gets DM list for members to attack and puts out their nation link
                        for index, member in enumerate(ping_list): 
                            link = f'https://politicsandwar.com/nation/id={nation_dict.get(member, "N/A")}'
                            user = client.get_user(member)
                            war_embed.add_field(name= f"__Attacker {index + 1}:__", value=f"[{user.display_name}]({link})", inline=True)
                            try:
                                await user.send(content = f"It's time to send in the elephants! Please check <#{channel.id}> for your **war assignment**. Thank You! P.S. Start with a **dogfight against their aircraft**, then **assassinate their spies** and post results in <#639621955795812362> :)")
                            except: 
                                await ctx.send(f'Could not DM {user.name} because they have disabled DMs with bots, please message them manually.')
                        def_ping_list = await coord_perms(defenders, channel, channel_name, ctx)
                        
                        ##Gets DM list for defending members and puts out their nation link
                        for index, member in enumerate(def_ping_list): 
                            link = f'https://politicsandwar.com/nation/id={nation_dict.get(member, "N/A")}'
                            user = client.get_user(member)
                            war_embed.add_field(name= f"__Defender {index + 1}:__", value=f"[{user.display_name}]({link})", inline=True)
                            try:
                                await user.send(content = f"Carthago is under siege! You've been attacked. Please check <#{channel.id}> to coordinate with your peers. Thank you!")
                            except: 
                                await ctx.send(f'Could not DM {user.name} because they have disabled DMs with bots, please message them manually.')
                        
                        war_embed.add_field(name="__Reminder__", value="1.) Make sure you have enough resources including food and uranium, ping gov if you need more\
                                \n 2.) Look over their military before going in and plan out the best move\
                                \n 3.) Talk and coordinate with fellow members, declare at the same time and help each other\
                                \n 4.) Again, start with a dogfight against their aircraft, then assassinate their spies and post results in <#639621955795812362>\
                                \n Good luck!", inline=False)
                        await channel.send(f'{ping(ping_list)}{ping(def_ping_list)}',embed = war_embed)

                        #Gets the permissions for the members set up and pings them

                    line += 1

                await ctx.send('Channels are finished being create, good luck in the wars to come!')

        elif api[0].lower() == 's':
            print('hi')
            with open(f'csv/{attachment.filename}') as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=',')
                update_dict()
                next(csv_reader)
                wars = []
                nations_to_api = set()
                #Goes through each row to create channels
                for row in csv_reader:
                    #Makes sure this is not the heading and it isn't an invalid row before creating a channel
                    if row[1] != '#ERROR!':
                        war = {}
                        war['target_name'] = row[1]
                        war['target_id'] = row[2]
                        war['attackers'] = [row[4], row[6], row[8]]
                        war['defenders'] = [row[11],row[13],row[15],row[17],row[19]]
                        wars.append(war)
                        nations_to_api.add(row[2])
                id_string = ",".join(nations_to_api)
                nations = requests.get(f'http://{sham_ip}:8080/nations/?key={sham_api_key}&limit=50&_id={id_string}&sort_key=score&sort_dir=-1&project={{"cities":1,"score":1,"soldiers":1,"tanks":1,"aircraft":1,"ships":1}}').json()
                war_db = {}
                for nation in nations:
                    war_db[nation['_id']] = nation
                
                for war in wars:
                    channel_name = war['target_name'].replace(' ', '-') + '-' + war['target_id']
                    mil_count = war_db[int(war['target_id'])]
                    #Creates the channel and edits the topic
                    channel = None
                    for category in list(filter(None, category_list_id)):
                        try:                             
                            channel = await ctx.guild.create_text_channel(channel_name, category = category, topic = f'War on https://politicsandwar.com/nation/id={war["target_id"]}')
                            break
                        except:
                            pass

                    war_embed = discord.Embed(title= f"⚔️ __Target: {' '.join(channel_name.split('-')[:-1])}__", 
                        description= f"Please declare {war_types[war_type]} war on https://politicsandwar.com/nation/id={war['target_id']}", color=0xcb2400,
                        url = f'https://politicsandwar.com/nation/war/declare/id={war["target_id"]}')


                    war_embed.add_field(name = '__Military Information:__', value = f'{channel_name.split("-")[0]} has {mil_count["soldiers"]} soldiers, {mil_count["tanks"]} tanks, {mil_count["aircraft"]} planes, and {mil_count["ships"]} ships', inline = False)

                    ping_list = await coord_perms(war['attackers'], channel, channel_name, ctx)
                    #Gets DM list for members to attack and puts out their nation link
                    for index, member in enumerate(ping_list): 
                        link = f'https://politicsandwar.com/nation/id={nation_dict.get(member, "N/A")}'
                        user = client.get_user(member)
                        war_embed.add_field(name= f"__Attacker {index + 1}:__", value=f"[{user.display_name}]({link})", inline=True)
                        try:
                            await user.send(content = f"It's time to send in the elephants! Please check <#{channel.id}> for your **war assignment**. Thank You! P.S. Start with a **dogfight against their aircraft**, then **assassinate their spies** and post results in <#639621955795812362> :)")
                        except: 
                            await ctx.send(f'Could not DM {user.name} because they have disabled DMs with bots, please message them manually.')
                    def_ping_list = await coord_perms(war['defenders'], channel, channel_name, ctx)
                    
                    ##Gets DM list for defending members and puts out their nation link
                    for index, member in enumerate(def_ping_list): 
                        link = f'https://politicsandwar.com/nation/id={nation_dict.get(member, "N/A")}'
                        user = client.get_user(member)
                        war_embed.add_field(name= f"__Defender {index + 1}:__", value=f"[{user.display_name}]({link})", inline=True)
                        try:
                            await user.send(content = f"Carthago is under siege! You've been attacked. Please check <#{channel.id}> to coordinate with your peers. Thank you!")
                        except: 
                            await ctx.send(f'Could not DM {user.name} because they have disabled DMs with bots, please message them manually.')
                    
                    war_embed.add_field(name="__Reminder__", value="1.) Make sure you have enough resources including food and uranium, ping gov if you need more\
                            \n 2.) Look over their military before going in and plan out the best move\
                            \n 3.) Talk and coordinate with fellow members, declare at the same time and help each other\
                            \n 4.) Again, start with a dogfight against their aircraft, then assassinate their spies and post results in <#639621955795812362>\
                            \n Good luck!", inline=False)
                    await channel.send(f'{ping(ping_list)}{ping(def_ping_list)}',embed = war_embed)

                    #Gets the permissions for the members set up and pings them


                await ctx.send('Channels are finished being create, good luck in the wars to come!')

    #If they don't have permission, tell them
    else:
        await ctx.send('You do not have permissions to create war channels')




"""
@client.command()
async def run_test_sheet(ctx):
    with open(f'csv/Blitz_Sheet_1_-_Sheet1.csv') as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        line = 0
        for row in csv_reader:
            if line != 0:
                print(f'{row[1]} is the target. Attacker 1 is {row[4]}, attacker 2 is {row[6]}, and attacker 3 is {row[8]}')
                target_name = row[1]
                target_id = row[2]
                attackers = [row[4], row[6], row[8]]
                defenders = [row[11],row[13],row[15],row[17],row[19]]

                category = discord.utils.get(ctx.guild.categories, name = 'War-Test')
                channel_name = target_name.replace(' ', '-') + '-' + target_id

                await ctx.guild.create_text_channel(channel_name, category = category)

                channel = discord.utils.get(ctx.guild.text_channels, name = channel_name.lower())

                await channel.edit(topic = f'War on {row[0]}')
                

                ping_list = await coord_perms(attackers, channel, channel_name, ctx)
                await channel.send(f'{ping(ping_list)}please declare war on {row[0]}\
                    \n 1.) Make sure you have enough resources including food and uranium, ping gov if you need more\
                    \n 2.) Look over their military before going in and plan out the best move\
                    \n 3.) Talk and coordinate with fellow members, declare at the same time and help each other\
                    \n Good luck!') 

                ping_list = await coord_perms(defenders, channel, channel_name, ctx)
                await channel.send(f'{ping(ping_list)}is/are defending against {target_name}') 

            line += 1
"""


@client.command()
async def create_chan(ctx, nation_link, war_type: Optional[int] = 0, reason: Optional[str] = "Carthago Counter", *members: Optional[discord.Member]):
    ''' 
    Creates a channel using nation link and the list of members to add to it

    :param nation_link: PnW link of nation that is the target
    :param reason: Reason (optional) for the war, you can leave it blank. "+" in place of spaces for reason.
    :param *members: Discord members to add to the channel
    '''
    category_list_id = []

    for category in category_list:
        category_list_id.append(discord.utils.get(ctx.guild.categories, name = category))

    category_list_id = list(filter(None, category_list_id))
    #Sees if the user has permissions to manage channels in the category and access bot

    #Checks if they have the permission to create these channels
    if any(category.permissions_for(ctx.author).manage_channels for category in category_list_id):

        #Makes sure that the nation_link is in the right format
        if(re.search(r'politicsandwar.com/nation/id=\d{1,7}', nation_link)):
            channel_name = get_pnw_name(nation_link).replace(' ', '-') + '-' + nation_link.split('=')[1]

            channel = None
            for category in category_list_id:
                print(category)
                try: 
                    channel = await ctx.guild.create_text_channel(channel_name, category = category, topic = f'War on {nation_link}')
                    print(channel)
                    break
                except:
                    pass
            update_dict()

            #Checks to make sure it is a war reason and not a member
            
            if re.match(r'<@\d{18}\>', reason) or re.match(r'<@!\d{18}\>', reason):
                id = int(reason.split('!')[1].split('>')[0])
                members += (client.get_user(id),)
                reason = ''
            #For loop to set permissions for members
            ping = ''
            for member in members:
                await channel.set_permissions(member, read_messages=True, send_messages=True)
                ping = ping + f'<@{member.id}> '
            
            #If it is a valid war reason, replace + with spaces
            if reason != '':
                reason = reason.replace('+', ' ')
                reason = f', war reason: {reason}'
  
            war_embed = discord.Embed(title= f"⚔️ __Target: {' '.join(channel_name.split('-')[:-1])}__", 
                description= f"Please declare {war_types[war_type]} war on on {nation_link}{reason}", color=0xcb2400,
                url = f'https://politicsandwar.com/nation/war/declare/id={nation_link.split("=")[1]}')


            mil_count = get_pnw_mil(f'https://politicsandwar.com/nation/id={nation_link.split("=")[1]}')

            war_embed.add_field(name = '__Military Information:__', value = f'{channel_name.split("-")[0]} has {mil_count["Soldiers"]} soldiers, {mil_count["Tanks"]} tanks, {mil_count["Planes"]} planes, and {mil_count["Ships"]} ships', inline = False)

            for index, member in enumerate(members): 
                link = f'https://politicsandwar.com/nation/id={nation_dict.get(member.id, "N/A")}'
                war_embed.add_field(name= f"__Attacker {index + 1}:__", value=f"[{member.display_name}]({link})", inline=True)
           
            war_embed.add_field(name="__Reminder__", value="1.) Make sure you have enough resources including food and uranium, ping gov if you need more\
                    \n 2.) Look over their military before going in and plan out the best move\
                    \n 3.) Talk and coordinate with fellow members, declare at the same time and help each other\
                    \n 4.) Again, start with a dogfight against their aircraft, then assassinate their spies and post results in <#639621955795812362>\
                    \n Good luck!", inline=False)

            await channel.send(f'{ping}',embed = war_embed)

        #Lets user know that nation_link is in wring format
        else:
            await ctx.send('Nation link format is wrong, must be politicsandwar.com/nation/id=xxxxx')
    #Tells them if they don't have the permission to create the channel
    else:
        await ctx.send('You do not have permissions to create war channels')




@client.command()
async def clear_expired(ctx):
    ''' 
    Deletes all the channels that have expired wars and have become usless

    '''
    category_list_id = []

    for category in category_list:
        category_list_id.append(discord.utils.get(ctx.guild.categories, name = category))

    category_list_id = list(filter(None, category_list_id))

    #Checks if they have the permission to create these channels
    if any(category.permissions_for(ctx.author).manage_channels for category in category_list_id):
        #Runs for loop thru all the channels to find obsolete ones
        for category in category_list_id:
            for channel in category.channels:
                active_war = False
                topic = channel.topic
                #try:
                nation_link = topic.split()[2]
                #Goes through a nation's wars and see if they still have wars with Carthago
                carth_and_nexus = ["Carthago", "House Stark", "Order of the White Rose"]
                for war in get_war_info(nation_link):
                    if war['Aggressor Alliance'] in carth_and_nexus or war['Defender Alliance'] in carth_and_nexus:
                        active_war = True
                        break

                if active_war == False:
                    await ctx.send(f'The war channel {channel.name} has been deleted.')
                    await channel.delete(reason="No existing Carthago wars with this nation")
                '''except AttributeError:
                    await ctx.send(f"The war channel {channel.name} does not have a topic, thus we are unable to determine if war has been expired")

                except: 
                    await ctx.send(f'Unexpected error has occured with {channel.name}, get piggy on it.')'''
    #Tells them if they don't have the permission to create the channel
    else:
        await ctx.send('You do not have permissions to create war channels')



@client.command()
@commands.cooldown(1, 60, commands.BucketType.channel)
async def war_info_full(ctx):
    '''
    Gets the all war information about the target nation

    '''
    message = await ctx.send('Gathering information... please wait a few moments')
    book = load_workbook('spreadsheet/War.xlsx')
    sheet = book.active #active means last opened sheet
    try:
        nation_link = ctx.channel.topic.split()[2]
        target_id = nation_link.split('=')[1]
        sheet[f'B3'] = get_pnw_name(nation_link)

        #clears all old cells
        for row in sheet['B3:S3']:
          for cell in row:
            cell.value = None

        for row in sheet['B6:S8']:
          for cell in row:
            cell.value = None

        for row in sheet['B10:S14']:
          for cell in row:
            cell.value = None

        #gets their basic military information and fills it in
        nation_info = req_info(nation_link)
        sheet[f'B3'] = nation_info['name']
        sheet[f'C3'] = nation_info['alliance']
        sheet[f'F3'] = nation_info['offensivewars']
        sheet[f'H3'] = nation_info['nationid']
        sheet[f'K3'] = nation_info['cities']
        sheet[f'L3'] = nation_info['soldiers']
        sheet[f'M3'] = nation_info['tanks']
        sheet[f'N3'] = nation_info['aircraft']
        sheet[f'O3'] = nation_info['ships']

        #Function to fill in war information for offensive and defensive wars@
        def war_fill(row, war, ID):
            sheet[f'D{row}'] = war['turns_left']

            #Basic control
            if war['ground_control'] == ID:
                sheet[f'P{row}'] = '✔'
            elif war['ground_control'] == target_id:
                sheet[f'P{row}'] = '✖'

            if war['air_superiority'] == ID:
                sheet[f'Q{row}'] = '✔'
            elif war['air_superiority'] == target_id:
                sheet[f'Q{row}'] = '✖'

            if war['blockade'] == ID:
                sheet[f'R{row}'] = '✔'
            elif war['blockade'] == target_id:
                sheet[f'R{row}'] = '✖'
            
            req = req_info(f'https://politicsandwar.com/nation/id={ID}')
            sheet[f'B{row}'] = req['leadername']
            sheet[f'K{row}'] = req['cities']
            sheet[f'E{row}'] = req['defensivewars']
            sheet[f'F{row}'] = req['offensivewars']
            sheet[f'L{row}'] = req["soldiers"]
            sheet[f'M{row}'] = req["tanks"]
            sheet[f'N{row}'] = req["aircraft"]
            sheet[f'O{row}'] = req["ships"]

        #Fils in offensive war unique attributes
        off_wars = get_all_war_info(nation_link, True)
        for index, war in enumerate(off_wars):
            row = index + 10
            sheet[f'C{row}'] = war['defender_alliance_name']
            sheet[f'G{row}'] = war['defender_military_action_points']
            sheet[f'H{row}'] = int(war['defender_resistance'])
            sheet[f'I{row}'] = int(war['aggressor_resistance'])
            sheet[f'J{row}'] = war['aggressor_military_action_points']
            war_fill(row, war, war['defender_id'])

        #Fills in defensive war unique attributes
        def_wars = get_all_war_info(nation_link, False)
        for index, war in enumerate(def_wars):
            row = index + 6
            sheet[f'C{row}'] = war['aggressor_alliance_name']
            sheet[f'G{row}'] = war['aggressor_military_action_points']
            sheet[f'H{row}'] = int(war['aggressor_resistance'])
            sheet[f'I{row}'] = int(war['defender_resistance'])
            sheet[f'J{row}'] = war['defender_military_action_points']
            war_fill(row, war, war['aggressor_id'])


        book.save('spreadsheet/War.xlsx')

        #excel2img.export_img("spreadsheet/War.xlsx","spreadsheet/image.png","War Scenario Sheet")
        #image = discord.File("spreadsheet/image.png", filename="war_screen.png")

        sheet = discord.File('spreadsheet/War.xlsx', filename="war_sheet.xlsx")
        await message.delete()
        await ctx.send(file = sheet)
    
    #Except block
    except AttributeError:
        await ctx.send(f"The war channel {ctx.channel.name} does not have the correct topic format. We can not grab information for the war.")

    except: 
        await ctx.send(f'Unexpected error has occured with {ctx.channel.name}, get piggy on it.')





@client.command()
@commands.cooldown(1, 30, commands.BucketType.user)
@commands.has_any_role(567389586934726677, 712071416505172090)
async def graph(ctx, type, *alliances): 
    '''
    Graphs specific information for alliances
    :param type: is the type of graph
    :param alliances: is the list of alliances to graph

    '''
    message = await ctx.send('Generating graph... please wait a few moments')

    #graphs histogram 
    if re.search('hist', type):
        alliances_data = []
        label = []

        #Makes it in the right format to be entered in url
        for name in alliances:
            label.append(name.replace('+', ' ').title())

        #Runs through alliances and sees the number of nations
        for ally in alliances:
            await message.edit(content = f'Gathering information from {ally.replace("+", " ").title()}')
            print(ally)
            res = requests.get(f'https://politicsandwar.com/index.php?id=15&keyword={ally}&cat=alliance&ob=score&od=DESC&maximum=15&minimum=0&search=Go&memberview=true')
            soup_data = BeautifulSoup(res.text, 'html.parser')
            data = soup_data.find(text = re.compile('Showing'))
            num_nations = float(data.split()[3])
            
            #Error handling for if the number of nations for the alliance is 0
            if num_nations == 0:
                await ctx.send(f'Could not find any nations in the alliance {ally.replace("+", " ").title()}, make sure it is spelled correctly')
                label.remove(ally)
                continue
            alliance_city_data = []

            #Grabs data for every nation in the alliance
            for nations in range(0, math.ceil(num_nations/50)):
                res = requests.get(f'https://politicsandwar.com/index.php?id=15&keyword={ally}&cat=alliance&ob=score&od=DESC&maximum={50*(nations+1)}&minimum={50*nations}&search=Go&memberview=true')
                soup_data = BeautifulSoup(res.text, 'html.parser')
                data = soup_data.find_all("td", attrs={"class": "right"}, text = re.compile(r'^[1-9]\d*$'))

                for city in data:
                    alliance_city_data.append(float(city.contents[0]))
                    #alliance_city_data[city.contents[0]] += 1
            alliance_city_data = np.array(alliance_city_data)
            alliances_data.append(alliance_city_data)

        #Creates graph
        fig, axes = plt.subplots(nrows=1, ncols=2)
        ax0, ax1 = axes.flatten()
        fig.set_figheight(10)
        fig.set_figwidth(20)

        #Comparison histogram
        n_bins = max(list(map(lambda x: np.amax(x), alliances_data))).astype(np.int64)+1
        if type == 'hist-static' or type == 'hist-static':
            n_bins = 45
            ax0.set_ylim(0, 100)
            ax1.set_ylim(0, 100)
        ax0.hist(alliances_data, range(n_bins+1), histtype='bar', label=label)
        ax0.legend(loc = 0)
        ax0.set_title('Comparison Graph')
        ax0.xaxis.set_ticks(np.arange(0, n_bins, 3))
        plt.setp(ax0.get_xticklabels(), rotation=30, horizontalalignment='right')

        #Stacked histogram
        ax1.hist(alliances_data, range(n_bins+1), histtype='bar', stacked=True)
        ax1.set_title('Stacked Histogram')
        ax1.xaxis.set_ticks(np.arange(0, n_bins, 1))
        plt.setp(ax1.get_xticklabels(), rotation=30, horizontalalignment='right')
        fig.savefig('graph.png', dpi = 300)

        #Sends file onto Discord
        pic = discord.File('graph.png', filename="graph.png")
        await message.delete()
        await ctx.send(file = pic)

    #Graphs for scatter plot with military informationS
    elif re.search('scat', type):
        milplot_array = np.array([['Alliance', 'Leader Name', 'Age', 'Cities', 'Soldiers Killed', 'Soldier Casualties'\
        , 'Tanks Killed', 'Tank Casualties', 'Planes Killed', 'Plane Casualties'\
        , 'Ships Killed', 'Ship Casualties', 'Infra Destroyed', 'Infra Lost', 'Money Looted']], dtype = object)

        html = []
        #Gets it in the right format for html
        for alliance in alliances: 
            html.append(alliance.replace(' ', '+'))

        alliances_data = []

        #Goes through each alliance and gets the nation links
        for ally in html:
            await message.edit(content = f'Gathering information from {ally.replace("+", " ").title()}')
            res = requests.get(f'https://politicsandwar.com/index.php?id=15&keyword={ally}&cat=alliance&ob=score&od=DESC&maximum=15&minimum=0&search=Go&memberview=true')
            soup_data = BeautifulSoup(res.text, 'html.parser')
            data = soup_data.find(text = re.compile('Showing'))
            num_nations = float(data.split()[3])
            #Error handling if alliance has no nations
            if num_nations == 0:
                await ctx.send(f'Could not find any nations in the alliance {ally.replace("+", " ").title()}, make sure it is spelled correctly')
                label.remove(ally)
                continue

            alliance_city_data = []
            #alliance_city_data = defaultdict(lambda: 0, alliance_city_data)

            #Goes to every nation in the alliance to grab military information
            for nations in range(0, math.ceil(num_nations/50)):
                res = requests.get(f'https://politicsandwar.com/index.php?id=15&keyword={ally}&cat=alliance&ob=score&od=DESC&maximum={50*(nations+1)}&minimum={50*nations}&search=Go&memberview=true')
                soup_data = BeautifulSoup(res.text, 'html.parser')
                data = soup_data.find_all("a", href=re.compile("politicsandwar.com/nation/id="))
                links = []

                for nation_link in data:
                    links.append(nation_link['href'])

                for nation_link in links:
                    mil = req_info(nation_link)

                    #Creates dataframe
                    member_info = [mil['alliance'], mil['leadername'], int(mil['daysold']), mil['cities'],\
                        int(mil['soldierskilled']), int(mil['soldiercasualties']),int(mil['tankskilled']),int(mil['tankcasualties']),\
                        int(mil['aircraftkilled']),int(mil['aircraftcasualties']),int(mil['shipskilled']),int(mil['shipcasualties']),\
                        float(mil['infdesttot']),int(mil['infraLost']),float(mil['moneyLooted'])]

                    temp_arr = np.array([member_info], dtype = object)

                    milplot_array = np.append(milplot_array, temp_arr, axis=0)
       

        await message.edit(content = f'Creating Graph')
        mil = pd.DataFrame(milplot_array[1:], range(len(milplot_array)-1), milplot_array[0])

        mil.to_excel('Excel_Sample.xlsx', sheet_name = 'Sheet1')
        df = pd.read_excel('Excel_Sample.xlsx', sheet_name = 'Sheet1')
        print(mil.head())

        stuff_to_graph = ['Money Looted', 'Infra Destroyed', 'Planes Killed']

        await message.delete()

        #Creates graph
        for image in stuff_to_graph:
            scattered0 = sns.lmplot(x = 'Age', y = image, data = df, hue = 'Alliance')
            ann(df, image)
            scattered0.savefig(f"Scatter Plot Of {image}.png")

            pic = discord.File(f"Scatter Plot Of {image}.png", filename=f"Scatter Plot Of {image}.png")
            await ctx.send(file = pic)
   #If type is something weird
    else:
        await ctx.send(f'We are unable to identify graph type: {type}. The valid graph types are hist/hist-static and scatter.')




@client.command()
async def add(ctx, type, reason = None, *nations): 
    ''' 
    Adds member to existing war channel

    :param type: Indicate if added members are attackers or defenders
    :param reason: The (optional) reason for war, you can leave this blank. Dashes "-" in place of spaces for reason.
    :param nations: Nation link or ID of list of members to add
    ''' 
    category_list_id = []

    for category in category_list:
        category_list_id.append(discord.utils.get(ctx.guild.categories, name = category))

    category_list_id = list(filter(None, category_list_id))

    #Checks if they have the permission to create these channels
    if any(category.permissions_for(ctx.author).manage_channels for category in category_list_id):
        reason = reason.replace('+', ' ')
        #Clears the war reason if it doesn't exist and in place is a nation link/ID
        if re.search(r'\d{1,7}', reason):
            nations += (reason,)
            reason = ''
        else:
            reason = f'The war reason is: {reason}'

        #If members to be added are attackers
        if type == 'attacker' or type == 'attackers' or type == 'atker' or type == 'atkers' or type == 'atk' or type == 'atks' or type == 'attack' or type == 'attacks':
            ping = await add_to_chan(ctx, nations)
            #If you couldn't find any of the members there is no point to ping, so this check exists
            if len(ping) != 0:
                ping_list = ' '.join([f'<@{member}>' for member in ping])
                await ctx.send(f'{ping_list} please read above and declare war on {ctx.channel.topic.split()[2]}. {reason}')

        #If members to be added are defenders
        elif type == 'defender' or type == 'defenders' or type == 'def' or type == 'defs' or type == 'defend' or type == 'defense':
            ping = await add_to_chan(ctx, nations)
            #If you couldn't find any of the members there is no point to ping, so this check exists
            if len(ping) != 0:
                ping_list = ' '.join([f'<@{member}>' for member in ping])
                target = " ".join(ctx.channel.name.split('-')[:-1])
                await ctx.send(f'{ping_list} is defending. Please coordinate with the other Carthago members here for the war against {target.title()}.')
       
        #If it is an invalid type of war
        else:
            await ctx.send('Invalid command format. Please do !add <atk/def> <nation link/id> <nation link/id> etc.')
    else:
        await ctx.send("Shoo shoo no perms")


@client.command()
@commands.has_any_role(567389586934726677, 712071416505172090)
async def find_targets_old(ctx, member, target_alliance, ground_max_percent = 120, ground_min_percent = 0, air_max_percent = 120, air_min_percent = 0): 
    ''' 
    Finds targets to attack in an enemy alliance for member

    :param member: The nation link or nation id of member
    :param target_alliance: Alliance to search for targets in
    :param ground_max_percent: Upper range (example 120%) of ground units enemy can have. Default is 120.
    :param ground_min_percent: Lower range (example 70%) of ground units enemy can have. Default is 40.
    :param air_max_percent: Upper range (example 120%) of planes enemy can have. Default is 120.
    :param air_min_percent: Lower range (example 70%) of planes enemy can have. Default is 40.
    ''' 
    member_info = {'leadername': '', 'score': 0, 'soldiers': 0, 'tanks': 0, 'aircraft': 0, 'ships': 0}
    #Sees if this is a nation link format
    if(re.search(r'politicsandwar.com/nation/id=\d{1,7}', member)):
        #Tries if this is valid nation link
        try:
            raw_member_info = ID_info(member.split('=')[1])
            member_info['leadername'] = raw_member_info['leadername']
            member_info['score'] = float(raw_member_info['score'])
            for key in ['soldiers', 'tanks', 'aircraft', 'ships']:
                member_info[key] = int(raw_member_info[key])
        except:
            await ctx.send('API key ran out or the member is not a valid nation ID or link')
            return
    #Sees if this is nation ID format
    elif(re.search(r'\d{1,7}', member)):
        #Tries if this is valid nation link
        try:
            raw_member_info = ID_info(member)
            member_info['leadername'] = raw_member_info['leadername']
            member_info['score'] = float(raw_member_info['score'])
            for key in ['soldiers', 'tanks', 'aircraft', 'ships']:
                member_info[key] = int(raw_member_info[key])
        except:
            await ctx.send('API key ran out or the member is not a valid nation ID or link')
            return
    #If invalid format
    else: 
        await ctx.send('The member is not a valid nation ID or link')
        return
    '''
    res = requests.get(f'https://politicsandwar.com/index.php?id=15&keyword={target_alliance}&cat=alliance&ob=score&od=DESC&maximum=15&minimum=0&search=Go&memberview=true&beige=true')
    soup_data = BeautifulSoup(res.text, 'html.parser')
    data = soup_data.find(text = re.compile('Showing'))
    num_nations = float(data.split()[3])
            
    #Error handling for if the number of nations for the alliance is 0
    if num_nations == 0:
        await ctx.send(f'Could not find any nations in the alliance {target_alliance.replace("+", " ").title()}, make sure it is spelled correctly')
        return

    loading_msg = await ctx.send('Generating a list of potential targets...')
    #Grabs data for every nation in the alliance
    alliance_nations_in_range = {}
    for nations in range(0, math.ceil(num_nations/50)):
        res = requests.get(f'https://politicsandwar.com/index.php?id=15&keyword={target_alliance}&cat=alliance&ob=score&od=DESC&maximum={50*(nations+1)}&minimum={50*nations}&search=Go&memberview=true&beige=true')
        soup_data = BeautifulSoup(res.text, 'html.parser')
        data = soup_data.find_all("td", attrs={"class": "right"}, text = re.compile(r'^[1-9]\d*$'))

        rows = soup_data.find("table", {'class': 'nationtable'}).find_all('tr')
        #Grabs nation score and nation ID for every nation
        for row in rows[1:]:
            cells = row.find_all('td')
            score = float(cells[6].text.replace(' ', '').replace(',', ''))
            try:
                slots = int(cells[6].find('img')['title'].split(' ')[0])
                if(member_info['score'] * 0.75 <= score <= member_info['score'] * 1.75 ):
                    alliance_nations_in_range[cells[1].find('a')['href'].split('=')[1]] = [cells[1].find('a').text, int(cells[5].text), score]
            except:
                pass

    '''
    target_alliance = target_alliance.lower()
    loading_msg = await ctx.send('Generating a list of potential targets...')
    update_spheres()
    print(spheres.keys())
    if(target_alliance.replace('+',' ') in spheres):
        nations = requests.get(f'http://{sham_ip}:8080/nations/?key={sham_api_key}&limit=1000&alliance={",".join(spheres[target_alliance.replace("+"," ")])}&defensivewars={{"$ne":3}}&color={{"$ne":"beige"}}&sort_key=score&sort_dir=-1&project={{"name":1,"cities":1,"score":1,"soldiers":1,"tanks":1,"aircraft":1,"ships":1}}').json()
    else:
        nations = requests.get(f'http://{sham_ip}:8080/nations/?key={sham_api_key}&limit=1000&alliance_name={target_alliance}&defensivewars={{"$ne":3}}&color={{"$ne":"beige"}}&sort_key=score&sort_dir=-1&project={{"name":1,"cities":1,"score":1,"soldiers":1,"tanks":1,"aircraft":1,"ships":1}}').json()
    
    alliance_nations_in_range = [nation for nation in nations if (member_info['score'] * 0.75 <= nation['score'] <= member_info['score'] * 1.75 )]

    target_embed = discord.Embed(title= f"🎯 __Potential Targets for {member_info['leadername']}__", 
        description = f'{member_info["leadername"]} has {member_info["soldiers"]} soldiers, {member_info["tanks"]} tanks, {member_info["aircraft"]} planes, and {member_info["ships"]} ships.')
    
    potential_targets = OrderedDict()
    for nation in alliance_nations_in_range:
        if (nation['soldiers'] + nation['tanks']*23 <= (member_info['soldiers'] + member_info['tanks'] * 23) * float(ground_max_percent)/100) and\
            (nation['soldiers'] + nation['tanks']*23 >= (member_info['soldiers'] + member_info['tanks'] * 23) * float(ground_min_percent)/100) and\
            nation['aircraft'] <= member_info['aircraft'] * float(air_max_percent)/100 and nation['aircraft'] >= member_info['aircraft'] * float(air_min_percent)/100:
            potential_targets[nation['_id']] = nation
            if not len(potential_targets) > 9:
                link = f"https://politicsandwar.com/nation/id={nation['_id']}"
                target_embed.add_field(name = '\u200b',
                value = f"[{nation['name']}]({link}) \n{nation['cities']} cities • {nation['soldiers']} soldiers • {nation['tanks']} tanks • {nation['aircraft']} planes • {nation['ships']} ships",
                inline = True)
    #Grabs military information for every nation in that list
    '''
    for nation_id in alliance_nations_in_range.keys():
        res = requests.post(f'https://politicsandwar.com/nation/id={nation_id}')
        soup_data = BeautifulSoup(res.text, 'html.parser')

        mil_values = []
        cells = soup_data.find("table", {'class': 'nationtable'}).find_all('td')
        
        #Finds the exact cells of the military information
        for cell in cells:
            if cell.find('br') != None and isinstance(cell.contents[0], str):
                mil_values.append(int(cell.contents[0].replace(',', '')))
        
        #Checks if it matches the parameters highlighted in command
        if (mil_values[0] + mil_values[1]*23 <= (member_info['soldiers'] + member_info['tanks'] * 23) * float(ground_max_percent)/100) and\
            (mil_values[0] + mil_values[1]*23 >= (member_info['soldiers'] + member_info['tanks'] * 23) * float(ground_min_percent)/100) and\
            mil_values[2] <= member_info['aircraft'] * float(air_max_percent)/100 and mil_values[2] >= member_info['aircraft'] * float(air_min_percent)/100:
            potential_targets[nation_id] = alliance_nations_in_range[nation_id] + mil_values
            if not len(potential_targets) > 9:
                link = f'https://politicsandwar.com/nation/id={nation_id}'
                target_embed.add_field(name = '\u200b',
                value = f'[{potential_targets[nation_id][0]}]({link}) \n{potential_targets[nation_id][1]} cities • {potential_targets[nation_id][3]} soldiers • {potential_targets[nation_id][4]} tanks • {potential_targets[nation_id][5]} planes • {potential_targets[nation_id][6]} ships',
                inline = True)
    '''
    await loading_msg.delete()
    target_embed.set_footer(text=f'Page 1/{int(math.ceil(len(potential_targets)/9))}')
    find_targets_msg = await ctx.send(embed = target_embed)
    
    #Only adds scrolling functions when more than 9 nations appear
    if len(potential_targets) > 9:
        await find_targets_msg.add_reaction('\u23ee')
        await find_targets_msg.add_reaction('\u25c0')
        await find_targets_msg.add_reaction('\u25b6')
        await find_targets_msg.add_reaction('\u23ed')


    i=0
    emoji=''
        #Keeps going until timeout error
    while True:
        #If it is to front error
        if emoji=='\u23ee':
            i=0
            target_embed.clear_fields()
            for nation in [x[1] for x in list(potential_targets.items())][0:9]:
                link = f"https://politicsandwar.com/nation/id={nation['_id']}"
                target_embed.add_field(name = '\u200b',
                value = f"[{nation['name']}]({link}) \n{nation['cities']} cities • {nation['soldiers']} soldiers • {nation['tanks']} tanks • {nation['aircraft']} planes • {nation['ships']} ships",
                inline = True)
            target_embed.set_footer(text=f'Page 1/{int(math.ceil(len(potential_targets)/9))}')
            await find_targets_msg.edit(embed=target_embed)
        #If left arrow
        if emoji=='\u25c0':
            if i>0:
                i-=1
                target_embed.clear_fields()
                for nation in [x[1] for x in list(potential_targets.items())][i*9:min(i*9+9,len(potential_targets))]:
                    link = f"https://politicsandwar.com/nation/id={nation['_id']}"
                    target_embed.add_field(name = '\u200b',
                    value = f"[{nation['name']}]({link}) \n{nation['cities']} cities • {nation['soldiers']} soldiers • {nation['tanks']} tanks • {nation['aircraft']} planes • {nation['ships']} ships",
                    inline = True)
                target_embed.set_footer(text=f'Page {i+1}/{int(math.ceil(len(potential_targets)/9))}')
                await find_targets_msg.edit(embed=target_embed)
        #If right arrow
        if emoji=='\u25b6':
            if i< int(len(potential_targets)/9):
                i+=1
                if float(i) == len(potential_targets)/9:
                    i -= 1
                target_embed.clear_fields()
                for nation in [x[1] for x in list(potential_targets.items())][i*9:min(i*9+9,len(potential_targets))]:
                    link = f"https://politicsandwar.com/nation/id={nation['_id']}"
                    target_embed.add_field(name = '\u200b',
                    value = f"[{nation['name']}]({link}) \n{nation['cities']} cities • {nation['soldiers']} soldiers • {nation['tanks']} tanks • {nation['aircraft']} planes • {nation['ships']} ships",
                    inline = True)
                target_embed.set_footer(text=f'Page {i+1}/{int(math.ceil(len(potential_targets)/9))}')
                await find_targets_msg.edit(embed=target_embed)
        #If to end arrow
        if emoji=='\u23ed':
            i=int(len(potential_targets)/9)
            if float(i) == len(potential_targets)/9:
                i -= 1
            target_embed.clear_fields()
            for nation in [x[1] for x in list(potential_targets.items())][i*9:min(i*9+9,len(potential_targets))]:
                link = f"https://politicsandwar.com/nation/id={nation['_id']}"
                target_embed.add_field(name = '\u200b',
                value = f"[{nation['name']}]({link}) \n{nation['cities']} cities • {nation['soldiers']} soldiers • {nation['tanks']} tanks • {nation['aircraft']} planes • {nation['ships']} ships",
                inline = True)
            target_embed.set_footer(text=f'Page {i+1}/{int(math.ceil(len(potential_targets)/9))}')
            await find_targets_msg.edit(embed=target_embed)

        #Checks for reactions
        try:
            res=await client.wait_for('reaction_add',timeout=90, check = lambda reaction, user: reaction.message.id == find_targets_msg.id)
            if str(res[1])!='War Pig#1807':
                emoji=str(res[0].emoji)
                await find_targets_msg.remove_reaction(res[0].emoji,res[1])
        #Breaks loop if ends
        except asyncio.TimeoutError:
            break

        #Removes reactions
        if str(res[1])!='War Pig#1807':
            emoji=str(res[0].emoji)
            await find_targets_msg.remove_reaction(res[0].emoji,res[1])
    '''
    #Keeps going until timeout error
    while True:
        #If it is to front error
        if emoji=='\u23ee':
            i=0
            target_embed.clear_fields()
            for nation_id in list(potential_targets.keys())[0:9]:
                link = f'https://politicsandwar.com/nation/id={nation_id}'
                target_embed.add_field(name = '\u200b',
                value = f'[{potential_targets[nation_id][0]}]({link}) \n{potential_targets[nation_id][1]} cities • {potential_targets[nation_id][3]} soldiers • {potential_targets[nation_id][4]} tanks • {potential_targets[nation_id][5]} planes • {potential_targets[nation_id][6]} ships',
                inline = True)
            target_embed.set_footer(text=f'Page 1/{int(math.ceil(len(potential_targets)/9))}')
            await find_targets_msg.edit(embed=target_embed)
        #If left arrow
        if emoji=='\u25c0':
            if i>0:
                i-=1
                target_embed.clear_fields()
                for nation_id in list(potential_targets.keys())[i*9:min(i*9+9,len(potential_targets))]:
                    link = f'https://politicsandwar.com/nation/id={nation_id}'
                    target_embed.add_field(name = '\u200b',
                    value = f'[{potential_targets[nation_id][0]}]({link}) \n{potential_targets[nation_id][1]} cities • {potential_targets[nation_id][3]} soldiers • {potential_targets[nation_id][4]} tanks • {potential_targets[nation_id][5]} planes • {potential_targets[nation_id][6]} ships',
                    inline = True)
                target_embed.set_footer(text=f'Page {i+1}/{int(math.ceil(len(potential_targets)/9))}')
                await find_targets_msg.edit(embed=target_embed)
        #If right arrow
        if emoji=='\u25b6':
            if i< int(len(potential_targets)/9):
                i+=1
                if float(i) == len(potential_targets)/9:
                    i -= 1
                target_embed.clear_fields()
                for nation_id in list(potential_targets.keys())[i*9:min(i*9+9,len(potential_targets))]:
                    link = f'https://politicsandwar.com/nation/id={nation_id}'
                    target_embed.add_field(name = '\u200b',
                    value = f'[{potential_targets[nation_id][0]}]({link}) \n{potential_targets[nation_id][1]} cities • {potential_targets[nation_id][3]} soldiers • {potential_targets[nation_id][4]} tanks • {potential_targets[nation_id][5]} planes • {potential_targets[nation_id][6]} ships',
                    inline = True)
                target_embed.set_footer(text=f'Page {i+1}/{int(math.ceil(len(potential_targets)/9))}')
                await find_targets_msg.edit(embed=target_embed)
        #If to end arrow
        if emoji=='\u23ed':
            i=int(len(potential_targets)/9)
            if float(i) == len(potential_targets)/9:
                i -= 1
            target_embed.clear_fields()
            for nation_id in list(potential_targets.keys())[i*9:min(i*9+9,len(potential_targets))]:
                link = f'https://politicsandwar.com/nation/id={nation_id}'
                target_embed.add_field(name = '\u200b',
                    value = f'[{potential_targets[nation_id][0]}]({link}) \n{potential_targets[nation_id][1]} cities • {potential_targets[nation_id][3]} soldiers • {potential_targets[nation_id][4]} tanks • {potential_targets[nation_id][5]} planes • {potential_targets[nation_id][6]} ships',
                    inline = True)
            target_embed.set_footer(text=f'Page {i+1}/{int(math.ceil(len(potential_targets)/9))}')
            await find_targets_msg.edit(embed=target_embed)

        #Checks for reactions
        try:
            res=await client.wait_for('reaction_add',timeout=90, check = lambda reaction, user: reaction.message.id == find_targets_msg.id)
            if str(res[1])!='War Pig#1807':
                emoji=str(res[0].emoji)
                await find_targets_msg.remove_reaction(res[0].emoji,res[1])
        #Breaks loop if ends
        except asyncio.TimeoutError:
            break

        #Removes reactions
        if str(res[1])!='War Pig#1807':
            emoji=str(res[0].emoji)
            await find_targets_msg.remove_reaction(res[0].emoji,res[1])
    '''
    await find_targets_msg.clear_reactions()


#CURRENT TESTING GROUND
@client.command()
@commands.has_any_role(567389586934726677, 712071416505172090)
async def find_counters(ctx, target, ground_max_percent = math.inf, ground_min_percent = 80, air_max_percent = math.inf, air_min_percent = 80): 
    await find_combined(ctx, "Counters", target, "carthago", 1/1.75, 1/0.75, ground_max_percent, ground_min_percent, air_max_percent, air_min_percent)

@client.command()
@commands.has_any_role(567389586934726677, 712071416505172090)
async def find_targets(ctx, member, target_alliance, ground_max_percent = 120, ground_min_percent = 0, air_max_percent = 120, air_min_percent = 0): 
    await find_combined(ctx, "Targets", member, target_alliance, 0.75, 1.75, ground_max_percent, ground_min_percent, air_max_percent, air_min_percent)


@client.command()
@commands.has_any_role(567389586934726677, 712071416505172090)
async def find_counters_old(ctx, target, ground_max_percent = math.inf, ground_min_percent = 80, air_max_percent = math.inf, air_min_percent = 80): 

    ''' 
    Finds counters for an enemy

    :param target: The nation link or nation id of target
    :param ground_max_percent: Upper range (example 120%) of ground units aa member can have. Default is infinity.
    :param ground_min_percent: Lower range (example 70%) of ground units aa member can have. Default is 80.
    :param air_max_percent: Upper range (example 120%) of planes aa member can have. Default is infinity.
    :param air_min_percent: Lower range (example 70%) of planes aa member can have. Default is 80.
    ''' 
    target_info = {'leadername': '', 'score': 0, 'soldiers': 0, 'tanks': 0, 'aircraft': 0, 'ships': 0}
    if(re.search(r'politicsandwar.com/nation/id=\d{1,7}', target)):
        try:
            raw_target_info = ID_info(target.split('=')[1])
            target_info['leadername'] = raw_target_info['leadername']
            target_info['score'] = float(raw_target_info['score'])
            target_info['slots'] = f'{raw_target_info["defensivewars"]}/3 slots'
            for key in ['soldiers', 'tanks', 'aircraft', 'ships']:
                target_info[key] = int(raw_target_info[key])
        except:
            await ctx.send('API key ran out or the member is not a valid nation ID or link')
            return
  
    elif(re.search(r'\d{1,7}', target)):
        try:
            raw_target_info = ID_info(target)
            target_info['leadername'] = raw_target_info['leadername']
            target_info['score'] = float(raw_target_info['score'])
            target_info['slots'] = f'{raw_target_info["defensivewars"]}/3 slots'
            for key in ['soldiers', 'tanks', 'aircraft', 'ships']:
                target_info[key] = int(raw_target_info[key])
        except:
            await ctx.send('API key ran out or the member is not a valid nation ID or link')
            return
 
    else: 
        await ctx.send('The member is not a valid nation ID or link')
        return

    loading_msg = await ctx.send('Generating a list of potential members to counter...')
    nations = requests.get(f'http://{sham_ip}:8080/nations/?key={sham_api_key}&limit=500&alliance_name=carthago&offensivewars={{"$ne":5}}&sort_key=score&sort_dir=-1&project={{"name":1,"cities":1,"score":1,"soldiers":1,"tanks":1,"aircraft":1,"ships":1}}').json()
    alliance_nations_in_range = [nation for nation in nations if (target_info['score'] * (1/1.75) <= nation['score'] <= target_info['score'] * (1/0.75) )]
    counter_embed = discord.Embed(title= f"🎯 __Potential Counter for {target_info['leadername']} ({target_info['slots']}):__", 
        description = f'{target_info["leadername"]} has {target_info["soldiers"]} soldiers, {target_info["tanks"]} tanks, {target_info["aircraft"]} planes, and {target_info["ships"]} ships.')
    

    potential_counters = OrderedDict()
    for nation in alliance_nations_in_range:
        if (nation['soldiers'] + nation['tanks']*23 <= max((target_info['soldiers'] + target_info['tanks'] * 23),0.01) * float(ground_max_percent)/100) and\
            (nation['soldiers'] + nation['tanks']*23 >= (target_info['soldiers'] + target_info['tanks'] * 23) * float(ground_min_percent)/100) and\
            nation['aircraft'] <= max(target_info['aircraft'],0.01) * float(air_max_percent)/100 and nation['aircraft'] >= target_info['aircraft'] * float(air_min_percent)/100:
            potential_counters[nation['_id']] = nation
            if not len(potential_counters) > 9:
                link = f"https://politicsandwar.com/nation/id={nation['_id']}"
                counter_embed.add_field(name = '\u200b',
                value = f"[{nation['name']}]({link}) \n{nation['cities']} cities • {nation['soldiers']} soldiers • {nation['tanks']} tanks • {nation['aircraft']} planes • {nation['ships']} ships",
                inline = True)
    '''
    res = requests.get(f'https://politicsandwar.com/index.php?id=15&keyword=Carthago&cat=alliance&ob=score&od=DESC&maximum=15&minimum=0&search=Go&memberview=true')
    soup_data = BeautifulSoup(res.text, 'html.parser')
    data = soup_data.find(text = re.compile('Showing'))
    num_nations = float(data.split()[3])
            
            #Grabs data for every nation in the alliance
    alliance_nations_in_range = {}
    for nations in range(0, math.ceil(num_nations/50)):
        res = requests.get(f'https://politicsandwar.com/index.php?id=15&keyword=Carthago&cat=alliance&ob=score&od=DESC&maximum={50*(nations+1)}&minimum={50*nations}&search=Go&memberview=true')
        soup_data = BeautifulSoup(res.text, 'html.parser')
        data = soup_data.find_all("td", attrs={"class": "right"}, text = re.compile(r'^[1-9]\d*$'))

        rows = soup_data.find("table", {'class': 'nationtable'}).find_all('tr')
        for row in rows[1:]:
            cells = row.find_all('td')
            score = float(cells[6].text.replace(' ', '').replace(',', ''))
            if(target_info['score'] * (1/1.75) <= score <= target_info['score'] * (1/0.75) ):
                alliance_nations_in_range[cells[1].find('a')['href'].split('=')[1]] = [cells[1].find('a').text, int(cells[5].text), score]

    counter_embed = discord.Embed(title= f"🎯 __Potential Counter for {target_info['leadername']} ({target_info['slots']}):__", 
        description = f'{target_info["leadername"]} has {target_info["soldiers"]} soldiers, {target_info["tanks"]} tanks, {target_info["aircraft"]} planes, and {target_info["ships"]} ships.')
   
    potential_counters = OrderedDict()
    for nation_id in alliance_nations_in_range.keys():
        res = requests.post(f'https://politicsandwar.com/nation/id={nation_id}')
        soup_data = BeautifulSoup(res.text, 'html.parser')

        mil_values = []
        cells = soup_data.find("table", {'class': 'nationtable'}).find_all('td')
        for cell in cells:
            if cell.find('br') != None and isinstance(cell.contents[0], str):
                mil_values.append(int(cell.contents[0].replace(',', '')))
    
        if (mil_values[0] + mil_values[1]*23 <= max((target_info['soldiers'] + target_info['tanks'] * 23),0.01) * float(ground_max_percent)/100) and\
            (mil_values[0] + mil_values[1]*23 >= (target_info['soldiers'] + target_info['tanks'] * 23) * float(ground_min_percent)/100) and\
            mil_values[2] <= max(target_info['aircraft'],0.01) * float(air_max_percent)/100 and mil_values[2] >= target_info['aircraft'] * float(air_min_percent)/100:
            potential_counters[nation_id] = alliance_nations_in_range[nation_id] + mil_values
            if not len(potential_counters) > 9:
                link = f'https://politicsandwar.com/nation/id={nation_id}'
                counter_embed.add_field(name = '\u200b',
                value = f'[{potential_counters[nation_id][0]}]({link}) \n{potential_counters[nation_id][1]} cities • {potential_counters[nation_id][3]} soldiers • {potential_counters[nation_id][4]} tanks • {potential_counters[nation_id][5]} planes • {potential_counters[nation_id][6]} ships',
                inline = True)
       '''
    counter_embed.set_footer(text=f'Page 1/{int(math.ceil(len(potential_counters)/9))}')

    await loading_msg.delete()
    find_counters_msg = await ctx.send(embed = counter_embed)
    if len(potential_counters) > 9:
        await find_counters_msg.add_reaction('\u23ee')
        await find_counters_msg.add_reaction('\u25c0')
        await find_counters_msg.add_reaction('\u25b6')
        await find_counters_msg.add_reaction('\u23ed')
    

    i=0
    emoji=''
    while True:
        if emoji=='\u23ee':
            i=0
            counter_embed.clear_fields()
            for nation in [x[1] for x in list(potential_counters.items())][0:9]:
                link = f"https://politicsandwar.com/nation/id={nation['_id']}"
                counter_embed.add_field(name = '\u200b',
                value = f"[{nation['name']}]({link}) \n{nation['cities']} cities • {nation['soldiers']} soldiers • {nation['tanks']} tanks • {nation['aircraft']} planes • {nation['ships']} ships",
                inline = True)
            counter_embed.set_footer(text=f'Page 1/{int(math.ceil(len(potential_counters)/9))}')
            await find_counters_msg.edit(embed=counter_embed)
        if emoji=='\u25c0':
            if i>0:
                i-=1
                counter_embed.clear_fields()
                for nation in [x[1] for x in list(potential_counters.items())][i*9:min(i*9+9,len(potential_counters))]:
                    link = f"https://politicsandwar.com/nation/id={nation['_id']}"
                    counter_embed.add_field(name = '\u200b',
                    value = f"[{nation['name']}]({link}) \n{nation['cities']} cities • {nation['soldiers']} soldiers • {nation['tanks']} tanks • {nation['aircraft']} planes • {nation['ships']} ships",
                    inline = True)
                counter_embed.set_footer(text=f'Page {i+1}/{int(math.ceil(len(potential_counters)/9))}')
                await find_counters_msg.edit(embed=counter_embed)
        if emoji=='\u25b6':
            if i< int(len(potential_counters)/9):
                i+=1
                if float(i) == len(potential_counters)/9:
                    i -= 1
                counter_embed.clear_fields()
                for nation in [x[1] for x in list(potential_counters.items())][i*9:min(i*9+9,len(potential_counters))]:
                    link = f"https://politicsandwar.com/nation/id={nation['_id']}"
                    counter_embed.add_field(name = '\u200b',
                    value = f"[{nation['name']}]({link}) \n{nation['cities']} cities • {nation['soldiers']} soldiers • {nation['tanks']} tanks • {nation['aircraft']} planes • {nation['ships']} ships",
                    inline = True)
                counter_embed.set_footer(text=f'Page {i+1}/{int(math.ceil(len(potential_counters)/9))}')
                await find_counters_msg.edit(embed=counter_embed)
        if emoji=='\u23ed':
            i=int(len(potential_counters)/9)
            if float(i) == len(potential_counters)/9:
                i -= 1
            counter_embed.clear_fields()
            for nation in [x[1] for x in list(potential_counters.items())][i*9:min(i*9+9,len(potential_counters))]:
                link = f"https://politicsandwar.com/nation/id={nation['_id']}"
                counter_embed.add_field(name = '\u200b',
                value = f"[{nation['name']}]({link}) \n{nation['cities']} cities • {nation['soldiers']} soldiers • {nation['tanks']} tanks • {nation['aircraft']} planes • {nation['ships']} ships",
                inline = True)
            counter_embed.set_footer(text=f'Page {i+1}/{int(math.ceil(len(potential_counters)/9))}')
            await find_counters_msg.edit(embed=counter_embed)
        try:
            res = await client.wait_for('reaction_add',timeout=90, check = lambda reaction, user: reaction.message.id == find_counters_msg.id)
            if str(res[1])!='War Pig#1807':
                emoji=str(res[0].emoji)
                await find_counters_msg.remove_reaction(res[0].emoji,res[1])
        except asyncio.TimeoutError:
            break

    '''
    while True:
        if emoji=='\u23ee':
            i=0
            counter_embed.clear_fields()
            for nation_id in list(potential_counters.keys())[0:9]:
                link = f'https://politicsandwar.com/nation/id={nation_id}'
                counter_embed.add_field(name = '\u200b',
                value = f'[{potential_counters[nation_id][0]}]({link}) \n{potential_counters[nation_id][1]} cities • {potential_counters[nation_id][3]} soldiers • {potential_counters[nation_id][4]} tanks • {potential_counters[nation_id][5]} planes • {potential_counters[nation_id][6]} ships',
                inline = True)
            counter_embed.set_footer(text=f'Page 1/{int(math.ceil(len(potential_counters)/9))}')
            await find_counters_msg.edit(embed=counter_embed)
        if emoji=='\u25c0':
            if i>0:
                i-=1
                counter_embed.clear_fields()
                for nation_id in list(potential_counters.keys())[i*9:min(i*9+9,len(potential_counters))]:
                    link = f'https://politicsandwar.com/nation/id={nation_id}'
                    counter_embed.add_field(name = '\u200b',
                    value = f'[{potential_counters[nation_id][0]}]({link}) \n{potential_counters[nation_id][1]} cities • {potential_counters[nation_id][3]} soldiers • {potential_counters[nation_id][4]} tanks • {potential_counters[nation_id][5]} planes • {potential_counters[nation_id][6]} ships',
                    inline = True)
                counter_embed.set_footer(text=f'Page {i+1}/{int(math.ceil(len(potential_counters)/9))}')
                await find_counters_msg.edit(embed=counter_embed)
        if emoji=='\u25b6':
            if i< int(len(potential_counters)/9):
                i+=1
                if float(i) == len(potential_counters)/9:
                    i -= 1
                counter_embed.clear_fields()
                for nation_id in list(potential_counters.keys())[i*9:min(i*9+9,len(potential_counters))]:
                    link = f'https://politicsandwar.com/nation/id={nation_id}'
                    counter_embed.add_field(name = '\u200b',
                    value = f'[{potential_counters[nation_id][0]}]({link}) \n{potential_counters[nation_id][1]} cities • {potential_counters[nation_id][3]} soldiers • {potential_counters[nation_id][4]} tanks • {potential_counters[nation_id][5]} planes • {potential_counters[nation_id][6]} ships',
                    inline = True)
                counter_embed.set_footer(text=f'Page {i+1}/{int(math.ceil(len(potential_counters)/9))}')
                await find_counters_msg.edit(embed=counter_embed)
        if emoji=='\u23ed':
            i=int(len(potential_counters)/9)
            if float(i) == len(potential_counters)/9:
                i -= 1
            counter_embed.clear_fields()
            for nation_id in list(potential_counters.keys())[i*9:min(i*9+9,len(potential_counters))]:
                link = f'https://politicsandwar.com/nation/id={nation_id}'
                counter_embed.add_field(name = '\u200b',
                    value = f'[{potential_counters[nation_id][0]}]({link}) \n{potential_counters[nation_id][1]} cities • {potential_counters[nation_id][3]} soldiers • {potential_counters[nation_id][4]} tanks • {potential_counters[nation_id][5]} planes • {potential_counters[nation_id][6]} ships',
                    inline = True)
            counter_embed.set_footer(text=f'Page {i+1}/{int(math.ceil(len(potential_counters)/9))}')
            await find_counters_msg.edit(embed=counter_embed)
        try:
            res = await client.wait_for('reaction_add',timeout=90, check = lambda reaction, user: reaction.message.id == find_counters_msg.id)
            if str(res[1])!='War Pig#1807':
                emoji=str(res[0].emoji)
                await find_counters_msg.remove_reaction(res[0].emoji,res[1])
        except asyncio.TimeoutError:
            break
    '''

    await find_counters_msg.clear_reactions()

@client.command()
async def sort_chans(ctx):
    category_list_id = []

    #Check if user has manage war chan perms aka are they milcom
    for category in category_list:
        category_list_id.append(discord.utils.get(ctx.guild.categories, name = category))

    category_list_id = list(filter(None, category_list_id))
    message = await ctx.send(f"Sorting channels now")

    #Checks if they have the permission to create these channels
    if any(category.permissions_for(ctx.author).manage_channels for category in category_list_id):
        for category in category_list_id:
            await message.edit(content= f"Sorting channels in {category.name}")
            channels = category.channels
            channels_copy = category.channels.copy()
            start_pos = channels[0].position
            await quick_sort(0, len(channels)-1, channels)
            #for channel in channels:
            print([channel.name for channel in channels])
            for index, channel in enumerate(channels):
                if(channel.name != category.channels[index].name):
                    await channel.edit(position = start_pos)
                start_pos += 1

    await message.edit(content= f"Finished sorting")

async def partition(start, end, array):
      
    # Initializing pivot's index to start
    pivot_index = start 
    pivot = array[pivot_index].name
      
    # This loop runs till start pointer crosses 
    # end pointer, and when it does we swap the
    # pivot with element on end pointer
    while start < end:
        # Increment the start pointer till it finds an 
        # element greater than  pivot 
        while start < len(array) and array[start].name <= pivot:
            start += 1
              
        # Decrement the end pointer till it finds an 
        # element less than pivot
        while array[end].name > pivot:
            end -= 1
          
        # If start and end have not crossed each other, 
        # swap the numbers on start and end
        if(start < end):
            '''
            start_pos = array[start].position
            end_pos = array[end].position
            await array[start].edit(position = end_pos)
            await array[end].edit(position = start_pos)'''
            array[start], array[end] = array[end], array[start]
      
    # Swap pivot element with element on end pointer.
    # This puts pivot on its correct sorted place.
    '''
    end_pos = array[start].position
    pivot_pos = array[pivot_index].position
    await array[pivot_index].edit(position = end_pos)
    await array[end].edit(position = pivot_pos)'''
    
    array[end], array[pivot_index] = array[pivot_index], array[end]
     
    # Returning end pointer to divide the array into 2
    return end
      
# The main function that implements QuickSort 
async def quick_sort(start, end, array):
    if (start < end):
          
        # p is partitioning index, array[p] 
        # is at right place
        p = await partition(start, end, array)
          
        # Sort elements before partition 
        # and after partition
        await quick_sort(start, p - 1, array)
        await quick_sort(p + 1, end, array)

@client.command()
async def piggy(ctx): 
    await ctx.send("Did you just assume <@236978935538122754>'s gender? Ew.")

@client.command()
async def mil_info(ctx): 
    nation_link = ctx.channel.topic.split()[2]
    nation_info = req_info(nation_link)
    await ctx.send(f"{nation_info['name']} has {nation_info['soldiers']} soldiers, {nation_info['tanks']} tanks, {nation_info['aircraft']} planes, and {nation_info['ships']} ships.")

@client.command()
async def active_wars(ctx): 
    war_json = requests.get(f'https://politicsandwar.com/api/wars/500&alliance_id=5049&key=69e9cc72114cd2').json()
    wars = war_json['wars']
    active_war_embed= discord.Embed(title= f"📓 Active Wars", 
        description = f'This is a list of active Carthago wars')
    for war in wars:
        if(war["status"] == "Active" or war["status"] == 'Defender Offered Peace' or war["status"] == 'Attacker Offered Peace'):
            name_atk = ID_info(f"{war['attackerID']}")['name']
            name_def = ID_info(f"{war['defenderID']}")['name']
            active_war_embed.add_field(name = war['warID'], value = f'{name_atk} vs {name_def}')

    await ctx.send(embed = active_war_embed)

async def find_combined(ctx, type, member, target_alliance, score_min, score_max, ground_max_percent, ground_min_percent, air_max_percent, air_min_percent):
    ''' 
    Finds targets to attack in an enemy alliance for member

    :param member: The nation link or nation id of member
    :param target_alliance: Alliance to search for targets in
    :param ground_max_percent: Upper range (example 120%) of ground units enemy can have. Default is 120.
    :param ground_min_percent: Lower range (example 70%) of ground units enemy can have. Default is 40.
    :param air_max_percent: Upper range (example 120%) of planes enemy can have. Default is 120.
    :param air_min_percent: Lower range (example 70%) of planes enemy can have. Default is 40.
    ''' 
    member_info = {'leadername': '', 'score': 0, 'soldiers': 0, 'tanks': 0, 'aircraft': 0, 'ships': 0}
    if(re.search(r'politicsandwar.com/nation/id=\d{1,7}', member)):
        try:
            raw_member_info = ID_info(member.split('=')[1])
            member_info['leadername'] = raw_member_info['leadername']
            member_info['score'] = float(raw_member_info['score'])
            member_info['slots'] = f'{raw_member_info["defensivewars"]}/3 slots'
            for key in ['soldiers', 'tanks', 'aircraft', 'ships']:
                member_info[key] = int(raw_member_info[key])
        except:
            await ctx.send('API key ran out or the member is not a valid nation ID or link')
            return
  
    #Sees if this is nation ID format
    elif(re.search(r'\d{1,7}', member)):
        #Tries if this is valid nation link
        try:
            raw_member_info = ID_info(member)
            member_info['leadername'] = raw_member_info['leadername']
            member_info['score'] = float(raw_member_info['score'])
            member_info['slots'] = f'{raw_member_info["defensivewars"]}/3 slots'
            for key in ['soldiers', 'tanks', 'aircraft', 'ships']:
                member_info[key] = int(raw_member_info[key])
        except:
            await ctx.send('API key ran out or the member is not a valid nation ID or link')
            return
    #If invalid format
    else: 
        await ctx.send('The member is not a valid nation ID or link')
        return
   
    target_alliance = target_alliance.lower()
    loading_msg = await ctx.send('Generating a list of potential targets...')
    update_spheres()
    slots =  "defensivewars"
    slot_num = 3
    if type == "Counters":
        slots = "offensivewars"
        slot_num = 5
    try:
        if(target_alliance.replace('+',' ') in spheres):
            #160.2.143.37 is the real
            nations = requests.get(f'http://{sham_ip}:8080/nations/?key={sham_api_key}&limit=1000&alliance={",".join(spheres[target_alliance.replace("+"," ")])}&{slots}={{"$ne":{slot_num}}}&color={{"$ne":"beige"}}&sort_key=score&sort_dir=-1&project={{"name":1,"cities":1,"score":1,"soldiers":1,"tanks":1,"aircraft":1,"ships":1}}').json()
        else:
            nations = requests.get(f'http://{sham_ip}:8080/nations/?key={sham_api_key}&limit=1000&alliance_name={target_alliance}&{slots}={{"$ne":{slot_num}}}&color={{"$ne":"beige"}}&sort_key=score&sort_dir=-1&project={{"name":1,"cities":1,"score":1,"soldiers":1,"tanks":1,"aircraft":1,"ships":1}}').json()
    except (requests.exceptions.ConnectionError):
        ctx.send("Trouble connecting to Shama's API.... Piggy and Shama have been notified")
        #PnW API
        '''
        res = requests.get(f'https://politicsandwar.com/index.php?id=15&keyword=Carthago&cat=alliance&ob=score&od=DESC&maximum=15&minimum=0&search=Go&memberview=true')
        soup_data = BeautifulSoup(res.text, 'html.parser')
        data = soup_data.find(text = re.compile('Showing'))
        num_nations = float(data.split()[3])
                
                #Grabs data for every nation in the alliance
        alliance_nations_in_range = {}
        for nations in range(0, math.ceil(num_nations/50)):
            res = requests.get(f'https://politicsandwar.com/index.php?id=15&keyword=Carthago&cat=alliance&ob=score&od=DESC&maximum={50*(nations+1)}&minimum={50*nations}&search=Go&memberview=true')
            soup_data = BeautifulSoup(res.text, 'html.parser')
            data = soup_data.find_all("td", attrs={"class": "right"}, text = re.compile(r'^[1-9]\d*$'))

            rows = soup_data.find("table", {'class': 'nationtable'}).find_all('tr')
            for row in rows[1:]:
                cells = row.find_all('td')
                score = float(cells[6].text.replace(' ', '').replace(',', ''))
                if(target_info['score'] * (1/1.75) <= score <= target_info['score'] * (1/0.75) ):
                    alliance_nations_in_range[cells[1].find('a')['href'].split('=')[1]] = [cells[1].find('a').text, int(cells[5].text), score]'''
        
        member = ctx.guild.get_member(236978935538122754)
        channel = await member.create_dm()
        await channel.send(f"Someone broke war pig: {sys.exc_info()[0]}, go annoy Shama")  

    alliance_nations_in_range = [nation for nation in nations if (member_info['score'] * score_min <= nation['score'] <= member_info['score'] * score_max)]

    target_embed = discord.Embed(title= f"🎯 __Potential {type} for {member_info['leadername'] } ({member_info['slots']}):__", 
        description = f'{member_info["leadername"]} has {member_info["soldiers"]} soldiers, {member_info["tanks"]} tanks, {member_info["aircraft"]} planes, and {member_info["ships"]} ships.')
         
    potential_targets = OrderedDict()
    for nation in alliance_nations_in_range:
        if (nation['soldiers'] + nation['tanks']*23 <= max((member_info['soldiers'] + member_info['tanks'] * 23),0.01) * float(ground_max_percent)/100) and\
            (nation['soldiers'] + nation['tanks']*23 >= (member_info['soldiers'] + member_info['tanks'] * 23) * float(ground_min_percent)/100) and\
            nation['aircraft'] <= max(member_info['aircraft'],0.01)  * float(air_max_percent)/100 and nation['aircraft'] >= member_info['aircraft'] * float(air_min_percent)/100:
            potential_targets[nation['_id']] = nation
            if not len(potential_targets) > 9:
                link = f"https://politicsandwar.com/nation/id={nation['_id']}"
                target_embed.add_field(name = '\u200b',
                value = f"[{nation['name']}]({link}) \n{nation['cities']} cities • {nation['soldiers']} soldiers • {nation['tanks']} tanks • {nation['aircraft']} planes • {nation['ships']} ships",
                inline = True)

    await loading_msg.delete()
    target_embed.set_footer(text=f'Page 1/{int(math.ceil(len(potential_targets)/9))}')
    find_targets_msg = await ctx.send(embed = target_embed)
    
    #Only adds scrolling functions when more than 9 nations appear
    if len(potential_targets) > 9:
        await find_targets_msg.add_reaction('\u23ee')
        await find_targets_msg.add_reaction('\u25c0')
        await find_targets_msg.add_reaction('\u25b6')
        await find_targets_msg.add_reaction('\u23ed')

    async def embed_creation(ctx, i, nation, potential_targets, target_embed):
        target_embed.clear_fields()
        for nation in [x[1] for x in list(potential_targets.items())][i*9:min(i*9+9,len(potential_targets))]:
            link = f"https://politicsandwar.com/nation/id={nation['_id']}"
            target_embed.add_field(name = '\u200b',
            value = f"[{nation['name']}]({link}) \n{nation['cities']} cities • {nation['soldiers']} soldiers • {nation['tanks']} tanks • {nation['aircraft']} planes • {nation['ships']} ships",
            inline = True)
        target_embed.set_footer(text=f'Page {i+1}/{int(math.ceil(len(potential_targets)/9))}')
        await find_targets_msg.edit(embed=target_embed)
   
    i=0
    emoji=''
        #Keeps going until timeout error
    while True:
        #If it is to front error
        if emoji=='\u23ee':
            i=0
            await embed_creation(ctx, i, nation, potential_targets, target_embed)
        #If left arrow
        if emoji=='\u25c0':
            if i>0:
                i-=1
                await embed_creation(ctx, i, nation, potential_targets, target_embed)
        #If right arrow
        if emoji=='\u25b6':
            if i< int(len(potential_targets)/9):
                i+=1
                if float(i) == len(potential_targets)/9:
                    i -= 1
                await embed_creation(ctx, i, nation, potential_targets,target_embed)
        #If to end arrow
        if emoji=='\u23ed':
            i=int(len(potential_targets)/9)
            if float(i) == len(potential_targets)/9:
                i -= 1
            await embed_creation(ctx, i, nation, potential_targets, target_embed)

        #Checks for reactions
        try:
            res=await client.wait_for('reaction_add',timeout=90, check = lambda reaction, user: reaction.message.id == find_targets_msg.id)
            if str(res[1])!='War Pig#1807':
                emoji=str(res[0].emoji)
                await find_targets_msg.remove_reaction(res[0].emoji,res[1])
        #Breaks loop if ends
        except asyncio.TimeoutError:
            break

        #Removes reactions
        if str(res[1])!='War Pig#1807':
            emoji=str(res[0].emoji)
            await find_targets_msg.remove_reaction(res[0].emoji,res[1])
  
    await find_targets_msg.clear_reactions()

async def coord_perms(members, channel, channel_name, ctx):
    ''' 
    Sets the permissions of the members by adding them to the coordination channel

    :param members: leader name of members to add to the channel
    :param channel: Channel object to add members to
    :param channel_name: Name of the channel
    :returns: String list of people to ping
    ''' 
    people_to_ping = []
    #Goes through every member in the list of members
    for member in members:
        #Only adds them if they're valid values
        if member != '#ERROR!' and member != '' and member != '#VALUE!' and member != "Loading...":
            print(member)
            discord_name = member_list(member)
            print(ctx.guild.get_member(discord_name))
            #Makes sure that it adds valid users
            if discord_name != '':
                await channel.set_permissions(ctx.guild.get_member(discord_name), read_messages=True, send_messages=True)
                people_to_ping.append(discord_name)
            #Tells gov to manually add users if they're invalid
            else:
                await ctx.send(f"Couldn't find the Discord for {member}, please add them manually to {channel_name}")
        #Stops the loop if invalid users come up because that means the end of the list
        else:
            return people_to_ping
            break
    #Returns it if there were no invalid users and the loop has reached its end
    else:
        return people_to_ping




async def add_to_chan(ctx, nations):
    ''' 
    Finds the members given the nation link and sets up permission for them

    :param nations: List of nations to be added
    :returns: Discord ID of list of people to ping
    ''' 
    members = []
    #For loop that goes through every nation in nations to find them
    for nation in nations:
        #If it is a link
        if len(nation.split("=")) > 1:
            nation_id = int(nation.split('=')[1])
        else:
            nation_id = int(nation)

        dis_id = requests.get(f'http://{sham_ip}:8080/discord/?key={sham_api_key}&_id={nation_id}').json()
        if(re.search(r'politicsandwar.com/nation/id=\d{1,7}', nation)) and len(dis_id) > 0:
            member = ctx.guild.get_member(int(dis_id[0]["DiscordID"]))
            await ctx.channel.set_permissions(member, read_messages=True, send_messages=True)
            members.append(int(dis_id[0]["DiscordID"]))

        #If it is an ID
        elif(re.match(r'\d{1,7}', nation)) and len(dis_id) > 0:
            member = ctx.guild.get_member(int(dis_id[0]["DiscordID"]))
            await ctx.channel.set_permissions(member, read_messages=True, send_messages=True)
            members.append(int(dis_id[0]["DiscordID"]))
            
        else:
            await ctx.send(f"Couldn't find member {nation}, either the member sheet is not updated or it is not a nation link/nation ID.")
    return members

@client.command()
@commands.has_any_role(567389586934726677, 712071416505172090)
async def help(ctx):
    help_embed = discord.Embed(title= f"📖 __List of Commands__", color=0xcb2400)
    help_embed.add_field(name = '!find_targets', value = f'Finds a list of targets in an alliance within range and military capabilities.\n Parameters\
        are <member nation id or link> <target alliance> <ground max %> <ground min %> <air max %> <air min %> (default is 120% max and 40% min, not neccessary to fill in)\
        \n__Example__: !find_targets 48730 The+Knights+Radiant 150 90 170 80 finds all TKR nations in range with 150-90% of my ground and 170-80% of my planes', inline = False)

    help_embed.add_field(name = '!find_counters', value = f'Finds a list of targets in Carthago within range and military capabilities to counter.\n Parameters\
        are <target nation id or link> <ground max %> <ground min %> <air max %> <air min %> (default is infinity% max and 80% min, not neccessary to fill in)\
        \n__Example__: !find_counters 48730 150 90 170 80 finds all Carthago nations in range with 150-90% of my ground and 170-80% of my planes', inline = False)
    
    category_list_id = []

    #Check if user has manage war chan perms aka are they milcom
    for category in category_list:
        category_list_id.append(discord.utils.get(ctx.guild.categories, name = category))

    category_list_id = list(filter(None, category_list_id))
    
    #Checks if they have the permission to create these channels
    if any(category.permissions_for(ctx.author).manage_channels for category in category_list_id):
        help_embed.add_field(name = '\u200b', value = '\u200b', inline = False)
        help_embed.add_field(name = '**__Milcom Specific Commands:__**', value = '\u200b', inline = False)
        help_embed.add_field(name = '!create_chan', value = f'Creates a channel for war.\n Parameters\
            are <target nation id or link> <Optional War Type 0, 1, 2> <Optional Counter Reason> <@member1> <@member2> etc\
            \n__Example__: !create_chan 48730 1 counter+for+nexus @Daveth#0674 @Kraмpus#0001 creates a channel telling them to declare ordinary war on Piglantia with reason of "counter for nexus"', inline = False)

        help_embed.add_field(name = '!bulk_create', value = f'Uses a CSV sheet to create a list of coordination channels.\n Parameter\
            is a CSV sheet using this format https://docs.google.com/spreadsheets/d/1Fo-wEUWkslONQE5tyIkLQ6Ubc3paPUMurx1SgQz8OFo/edit?usp=sharing', inline = False)

        help_embed.add_field(name = '!add', value = f'Adds attacker or defender to channel.\n Parameters are\
           <atk or def> <member id/link> \n Example: !add atk 48730 adds Piglantia to channel as attacker', inline = False)

        help_embed.add_field(name = '!war_info', value = f'Use in war channel to get excel sheet of MAPs and resistance of all wars target is in.', inline = False)

        help_embed.add_field(name = '!clear_expired', value = f'Deletes all war channels which no longer have active Carthago wars', inline = False)

        help_embed.add_field(name = '!graph', value = f'Creates graphs of the alliances\n Parameters\
            are <scat or histk> <alliance 1> <alliance 2> (use + to seperate spaces)\
            \n__Example__: !graph scat Carthago House+Stark creates a scatter graph of military kills and highlights outliers to be used as prio targets\n\
            !graph hist Carthago returns histogram of city count of Carthago', inline = False)


    await ctx.send(embed = help_embed)


@client.command()
@commands.cooldown(1, 30, commands.BucketType.channel)
async def war_info(ctx):
    try:
        nation_link = ctx.channel.topic.split()[2]
        await war_info_combined(ctx, nation_link)
    except:
        await ctx.send("Something went wrong :( likely with the channel set up. Go grab Piggu.")

@client.command()
@commands.has_any_role(567389586934726677, 712071416505172090)
@commands.cooldown(1, 30, commands.BucketType.channel)
async def wars(ctx, member):
    if(re.search(r'politicsandwar.com/nation/id=\d{1,7}', member)):
        await war_info_combined(ctx, member)

    elif(re.search(r'<@\d{16,19}>', member) or re.search(r'<@!\d{16,19}>', member)):
        try:
            disc_id = re.sub("[^0-9]", "", member)
            nation_id = requests.get(f'http://{sham_ip}:8080/discord/?key={sham_api_key}&DiscordID={disc_id}').json()[0]['_id']
            await war_info_combined(ctx, f'politicsandwar.com/nation/id={nation_id}')
        except:
            await ctx.send("Could not find member in database")
    elif(re.search(r'\d{1,7}', member)):
        await war_info_combined(ctx, f'politicsandwar.com/nation/id={member}')
    else:
        await ctx.send("Invalid link/ID")

async def war_info_combined(ctx, nation_link):
    nation_info = req_info(nation_link)
    if('error' not in nation_info):
        war_info_embed = discord.Embed(title= f"{nation_info['name']} ({nation_info['cities']}) [{nation_info['alliance']}] ", color=0x6AA84F)
        war_info_embed.add_field(name = 'War Info', value = f"The target has {nation_info['soldiers']} soldiers, {nation_info['tanks']} tanks, {nation_info['aircraft']} aircraft, and {nation_info['ships']} ships", inline = False)


        await ctx.send(embed = war_info_embed)

        #Fils in offensive war unique attributes
        off_embed = discord.Embed(title= f"Offensive Wars", color=0xcb2400)
        def_embed = discord.Embed(title= f"Defensive Wars", color=0x06FFFF)
        off_wars = get_all_war_info(nation_link, True)
        for war in off_wars:
            req = req_info(f"https://politicsandwar.com/nation/id={war['defender_id']}")
            def_info = ID_info(war['defender_id'])
            off_embed.add_field(name = f"{req['leadername']} ({req['cities']}) [{war['defender_alliance_name']}] https://politicsandwar.com/nation/id={war['defender_id']}", 
            value = f"{nation_info['name']} MAP: {war['aggressor_military_action_points']} | Resis: {war['aggressor_resistance']}\n{req['leadername']} MAP: {war['defender_military_action_points']} | Resis: {war['defender_resistance']}\n{def_info['soldiers']} soldiers | {def_info['tanks']} tanks | {def_info['aircraft']} aircraft | {def_info['ships']} ship", inline = False)

        #Fills in defensive war unique attributes
        def_wars = get_all_war_info(nation_link, False)
        for war in def_wars:
            req = req_info(f"https://politicsandwar.com/nation/id={war['aggressor_id']}")
            agg_info = ID_info(war['aggressor_id'])
            def_embed.add_field(name = f"{req['leadername']} ({req['cities']}) [{war['aggressor_alliance_name']}] https://politicsandwar.com/nation/id={war['aggressor_id']}", 
            value = f"{req['leadername']} MAP: {war['aggressor_military_action_points']} | Resis: {war['aggressor_resistance']}\n{nation_info['name']} MAP: {war['defender_military_action_points']} | Resis: {war['defender_resistance']}\n{agg_info['soldiers']} soldiers | {agg_info['tanks']} tanks | {agg_info['aircraft']} aircraft | {agg_info['ships']} ship", inline = False)

        await ctx.send(embed = off_embed)
        await ctx.send(embed = def_embed)
    else:
        await ctx.send("Invalid link/ID")

#THIS NEEDS TO BE UPDATED
@client.event
async def on_member_update(before, after):
    if len(before.roles) < len(after.roles):
        # The user has gained a new role, so lets find out which one
        newRole = next(role for role in after.roles if role not in before.roles)

        if newRole.name == "Citizen" or newRole.name == "Trainee":
            if after.id not in nation_dict:
                res = requests.get(f'http://{sham_ip}:8080/discord/?key={sham_api_key}&DiscordID={after.id}').json()[0]
                warmembergsheet.append_row([res['leader'], int(res['_id']), str(after), int(after.id), 0])
                update_dict()
'''
@client.command()
async def table(ctx, type, *aaorsphere): 
    message = await ctx.send('Gathering information... please wait a few moments')
    book = load_workbook('spreadsheet/CooperPooper.xlsx')
    sheet = book.active #active means last opened sheet
    try:
        for row in sheet['B1:K16']:
          for cell in row:
            cell.value = None

        if(type == 'alliances'):
            label = []

            #Makes it in the right format to be entered in url
            for name in alliances:
                label.append(name.replace('+', ' ').title())

            #Runs through alliances and sees the number of nations
            for i, ally in enumerated(alliances):
                await message.edit(content = f'Gathering information from {ally.replace("+", " ").title()}')
                print(ally)
                res = requests.get(f'https://politicsandwar.com/index.php?id=15&keyword={ally}&cat=alliance&ob=score&od=DESC&maximum=15&minimum=0&search=Go&memberview=true')
                soup_data = BeautifulSoup(res.text, 'html.parser')
                data = soup_data.find(text = re.compile('Showing'))
                num_nations = float(data.split()[3])
                
                #Error handling for if the number of nations for the alliance is 0
                if num_nations == 0:
                    await ctx.send(f'Could not find any nations in the alliance {ally.replace("+", " ").title()}, make sure it is spelled correctly')
                    label.remove(ally)
                    continue
                alliance_city_data = {'1-8': 0, 
                                    '8-10': 0,
                                    '10-12': 0,
                                    '13-14': 0,
                                    '15-16': 0,
                                    '17-18': 0,
                                    '19-20': 0,
                                    '21-23': 0,
                                    '24-26': 0,
                                    '27-30': 0,
                                    '31-34': 0,
                                    '35-60':0}

                #Grabs data for every nation in the alliance
                for nations in range(0, math.ceil(num_nations/50)):
                    res = requests.get(f'https://politicsandwar.com/index.php?id=15&keyword={ally}&cat=alliance&ob=score&od=DESC&maximum={50*(nations+1)}&minimum={50*nations}&search=Go&memberview=true')
                    soup_data = BeautifulSoup(res.text, 'html.parser')
                    data = soup_data.find_all("td", attrs={"class": "right"}, text = re.compile(r'^[1-9]\d*$'))

                    for city in data:
                        for key in alliance_city_data:
                            cityRange = list(map(int, key.split('-')))
                            if int(city.contents[0]) in range(key.split('-')[0],key.split('-')[0]+1):
                                alliance_city_data[key]+=1
                                break
            
  

        if(type == 'spheres'):
            pass
    except:
        await ctx.send('hi')
'''

'''
@client.command()
async def find_members(ctx):
    member_dis = ['Kraмpus','Kebab','Daveth','Tumatauenga','michaelborg','Grandmaster Bee',
    'CoolPerson125','bmber','Cyr_0nos','Asierith','KaiserLuch','Rag',
    'Tamasith','velium','Rohan Nair','Billy','Petko | Kosi |' 
    'Estival','Hans','ᴛʜᴇᴋɪᴅᴍᴀᴅᴇɪᴛ!','Duckman','Yuri Molotov','Cyrus' ,
    'NapoleonIII','Zeannon','Vedanshu','Gladiator Greyman','JayC',
    'GM (Noah)','NotCool','CthulhuTDOW','Strett',
    'Gallant Aegis','DyNasty','Waldo','Kingdracula','Leigh','Gust','Can','Madder Red#']

    member_ids = list(map(lambda member: discord.utils.get(ctx.guild.members, name = member).id if discord.utils.get(ctx.guild.members, name = member)!= None else None,member_dis))
    print(member_ids)'''




def member_list(leader_name):
    '''
    Find the discord ID of the member given their leader name

    :returns: The Discord ID of the member
    '''

    member_disc = member_dict.get(leader_name)
    if member_disc != None:
        return member_disc

    else:
        shama_db = requests.get(f'http://{sham_ip}:8080/discord/?key={sham_api_key}&leader={leader_name.replace(" ", "+")}').json()
        if len(shama_db) > 0:
            nation_dict[int(shama_db[0]["DiscordID"])] = shama_db[0]["_id"]
            return int(shama_db[0]["DiscordID"])
        else:
            return ""





def ping(ping_list):
    '''
    Creates a string in which the bot can then ping all the members

    :returns: The string in which the bot can ping all the members to declare war
    '''
    ping_str = ''
    for member in ping_list:
        ping_str += f'<@{str(member)}> '
    return ping_str



def ann(df, value):
    '''
    Annotates the graph with the outliers
    :param df: is the dataframe
    :param value: Is the value to find outliers in

    '''
    slope, intercept, r_value, p_value, std_err = stats.linregress(df['Age'],df[value])
    std_resid = math.sqrt((sum((df[value] - slope * df['Age'])**2))/(len(df.index)-2))
    outliers = df[df[value] > (slope * df['Age'] + 0.8 * std_resid)]
    plt.ylim(0, None)
    plt.title(f'Scatter Plot of Nation Age Vs {value}')
    for row in outliers.iterrows():
        r = row[1]
        plt.gca().annotate(f'{r["Leader Name"]}, {r["Cities"]}', xy=(r['Age'], r[value]), xytext=(2,2) , textcoords ="offset points", )



def update_dict():
    global member_dict
    global nation_dict
    '''
    member_names = gsheet.col_values(1)[1:]
    nation_id = [int(nation) for nation in gsheet.col_values(2)[1:]]
    dis_id = [int(disc) for disc in gsheet.col_values(4)[1:]]
    member_dict = dict(zip(member_names, dis_id))
    nation_dict = dict(zip(dis_id, nation_id))'''

    membership_db = requests.get(f'http://{sham_ip}:8080/discord/?key={sham_api_key}&alliance=5049').json()
    member_names = [member['leader'] for member in membership_db]
    dis_id = [int(member['DiscordID']) for member in membership_db]
    member_dict = dict(zip(member_names, dis_id))
    nation_dict = dict(zip(dis_id, nation_id))

def update_spheres():
    global spheres
    sphere_names = [sphere.lower() for sphere in wargsheet.col_values(1)[1:]]
    print(sphere_names)
    sphere_alliances = [sphere.split(',') for sphere in wargsheet.col_values(3)[1:]]
    spheres = dict(zip(sphere_names, sphere_alliances))





@create_chan.error
async def create_chan_error(ctx, error):
    '''
    Function to catch any errors in create_chan
    '''
    #If the member given is invalid
    if isinstance(error, commands.BadArgument):
        await ctx.send('I could not find the member(s)')
    #If the user forgets to input in members or the nation link
    if isinstance(error, commands.MissingRequiredArgument):
        await ctx.send('Missing required argument, the correct format is:\
            !create_chan [nation_link] @member @member ... etc')
    else:
        await ctx.send("API error")

#If the command is on cooldown
@war_info.error
async def war_info_error(ctx, error):
    if isinstance(error, commands.CommandOnCooldown):
        await ctx.send(f'The command is on cooldown for this channel, please try again in {error.retry_after:.3g} seconds')

@wars.error
async def wars(ctx, error):
    if isinstance(error, commands.CommandOnCooldown):
        await ctx.send(f'The command is on cooldown for this channel, please try again in {error.retry_after:.3g} seconds')

@graph.error
async def graph_error(ctx, error):
    if isinstance(error, commands.CommandOnCooldown):
        await ctx.send(f'The command is on cooldown for you, please try again in {error.retry_after:.3g} seconds')


token = os.environ.get('BOT_TOKEN')
client.run(f'{token}')
